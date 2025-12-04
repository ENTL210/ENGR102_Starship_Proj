import os
import time
import json 
import random
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from pathlib import Path
import datetime


def generate_answer_keys(export_directory):
    print("-" * 75)
    print((" " * 22) + "[2] Generating Answer Keys" + (" " * 22))
    print("-" * 75)
    
    path = Path(export_directory)
    datasets_count = 0
    result = []
    
    # Getting numbers of datapoints...
    datapoint_counts = len(generate_time_stamps())
    
    # Check if there is an dataset...
    for item in path.iterdir():
        if item.is_dir() and "Dataset" in item.name:
            datasets_count += 1
    
    # If there is no dataset...
    if datasets_count == 0:
        print("\n       There is no dataset in this directory.")
        print("\n       An answer key can't be generated. Exiting in 3 seconds...")
        print("-" * 75)
        for i in range(0,3):
            print(f"          Exiting program in {3-i}...") 
            time.sleep(1)
        print("-" * 75)
        exit()
    else: 
        print(f"\n      Numbers of datasets found: {datasets_count}")
        print("\n      Begin to loop through each dataset...")
    
    for item in path.iterdir():
        if item.is_dir() and "Dataset" in item.name:
            print(f"\n      {item.name}")
            print(f"\n           Counts of data points: {datapoint_counts}")
            
            # Calculating avg and max of temperature...
            temperature_file_name = "temps_in_C.xlsx"
            temperature_file_path = os.path.join(item, temperature_file_name)
            print(f"\n           Calculating avg & max temperatures:")
            # Loading the workbook...
            wb = load_workbook(temperature_file_path)
            ws = wb.active
            temperature_values = []
            for row_num in range(2, datapoint_counts + 2):
                cell_value = ws.cell(row_num, 2).value
                temperature_values.append(cell_value)
            avg_temperature_value = round(sum(temperature_values) / len(temperature_values), 2)
            max_temperature = round(max(temperature_values), 2)
            print(f"              Avg Temperature in C: {avg_temperature_value}°C")
            print(f"              Max Temperature in C: {max_temperature}°C")
            
            # Calculating avg and max of UV index...
            uv_file_name = "uv_index.xlsx"
            uv_file_path = os.path.join(item, uv_file_name)
            print(f"\n           Calculating avg & max UV:")
            # Loading the workbook...
            wb = load_workbook(uv_file_path)
            ws = wb.active
            uv_values = []
            for row_num in range(2, datapoint_counts + 2):
                cell_value = ws.cell(row_num, 2).value
                uv_values.append(cell_value)
            avg_uv_values = round(sum(uv_values) / len(uv_values), 2)
            max_uv_index = round(max(uv_values), 2)
            print(f"              Avg UV Index: {avg_uv_values}")
            print(f"              Max UV Index: {max_uv_index}")
            
            # Sum standard power input...
            standard_power_input_file_name = "power_input_std.xlsx"
            standard_power_input_file_path = os.path.join(item, standard_power_input_file_name)
            print(f"\n           Summing the Total Power the Standard Model Received:")
            # Loading the workbook...
            wb = load_workbook(standard_power_input_file_path)
            ws = wb.active
            standard_power_input_values = []
            for row_num in range(2, datapoint_counts + 2):
                cell_value = ws.cell(row_num, 2).value
                standard_power_input_values.append(cell_value)
            sum_of_standard_power_input = round(sum(standard_power_input_values), 2)
            print(f"              Sum of the Power the Standard Model Received: {sum_of_standard_power_input} Watts")
            
            # Sum PV power input...
            pv_power_input_file_name = "power_input_pv.xlsx"
            pv_power_input_file_path = os.path.join(item, pv_power_input_file_name)
            print(f"\n           Summing the Total Power Proposed Model received:")
            # Loading the workbook...
            wb = load_workbook(pv_power_input_file_path)
            ws = wb.active
            pv_power_input_values = []
            for row_num in range(2, datapoint_counts + 2):
                cell_value = ws.cell(row_num, 2).value
                pv_power_input_values.append(cell_value)
            sum_of_pv_power_input = round(sum(pv_power_input_values), 2)
            print(f"              Sum of the Power Proposed Model received: {sum_of_pv_power_input} Watts")
            
            # Sum standard power consumptions...
            standard_power_consumptions_file_name = "power_consumptions_std.xlsx"
            standard_power_consumptions_file_path = os.path.join(item, standard_power_consumptions_file_name)
            print(f"\n           Summing the Power the Standard Model consumed:")
            # Loading the workbook...
            wb = load_workbook(standard_power_consumptions_file_path)
            ws = wb.active
            standard_power_consumptions_values = []
            for row_num in range(2, datapoint_counts + 2):
                cell_value = ws.cell(row_num, 2).value
                standard_power_consumptions_values.append(cell_value)
            sum_of_standard_power_consumptons = round(sum(standard_power_consumptions_values), 2)
            print(f"              Sum of the Power the Standard Model consumed: {sum_of_standard_power_consumptons} Watts")
            
            # Sum PV power consumptions...
            pv_power_consumptions_file_name = "power_consumptions_pv.xlsx"
            pv_power_consumptions_file_path = os.path.join(item, pv_power_consumptions_file_name)
            print(f"\n           Summing the Power the Standard Model consumed:")
            # Loading the workbook...
            wb = load_workbook(pv_power_consumptions_file_path)
            ws = wb.active
            pv_power_consumptions_values = []
            for row_num in range(2, datapoint_counts + 2):
                cell_value = ws.cell(row_num, 2).value
                pv_power_consumptions_values.append(cell_value)
            sum_of_pv_power_consumptons = round(sum(pv_power_consumptions_values), 2)
            print(f"              Sum of the Power the Proposed Model consumed: {sum_of_pv_power_consumptons} Watts")
            
            # Calculating the cost of power consumptions...
            print(f"\n           Calculating the cost of power consumptions...")
            pv_avg_power_consumptions = sum(pv_power_consumptions_values) / len(pv_power_consumptions_values)
            standard_avg_power_consumptions = sum(standard_power_consumptions_values) / len(standard_power_consumptions_values)
            cost_of_power_consumptions = round(pv_avg_power_consumptions - standard_avg_power_consumptions, 2)
            print(f"              The Additional Power Consumed by the Proposed Model: {cost_of_power_consumptions} Watts")
            
            # Calculating the cost in payload volume
            print(f"\n           Calculating the cost in payload volume (m^3)")
            standard_payload_volume_file_name = "payload_volume_std.xlsx"
            standard_payload_volume_file_path = os.path.join(item, standard_payload_volume_file_name)
            wb = load_workbook(standard_payload_volume_file_path)
            ws = wb.active
            standard_payload_volume_values = []
            for row_num in range(2, datapoint_counts + 2):
                cell_value = ws.cell(row_num, 2).value
                standard_payload_volume_values.append(cell_value)
            avg_standard_payload_volume = sum(standard_payload_volume_values) / len(standard_payload_volume_values)
            
            pv_payload_volume_file_name = "payload_volume_pv.xlsx"
            pv_payload_volume_file_path = os.path.join(item, pv_payload_volume_file_name)
            wb = load_workbook(pv_payload_volume_file_path)
            ws = wb.active
            pv_payload_volume_values = []
            for row_num in range(2, datapoint_counts + 2):
                cell_value = ws.cell(row_num, 2).value
                pv_payload_volume_values.append(cell_value)
            avg_pv_payload_volume = sum(pv_payload_volume_values) / len(pv_payload_volume_values)
            cost_of_payload_volume = round((avg_standard_payload_volume - avg_pv_payload_volume) / avg_standard_payload_volume * 100, 2)
            print(f"              The Proposed Model reduced the payload volume (m^3) by {cost_of_payload_volume}%")
            
            # Calculating the cost in payload mass
            print(f"\n           Calculating the cost in payload mass (kg)")
            standard_payload_mass_file_name = "payload_in_kg_std.xlsx"
            standard_payload_mass_file_path = os.path.join(item, standard_payload_mass_file_name)
            wb = load_workbook(standard_payload_mass_file_path)
            ws = wb.active
            standard_payload_mass_values = []
            for row_num in range(2, datapoint_counts + 2):
                cell_value = ws.cell(row_num, 2).value
                standard_payload_mass_values.append(cell_value)
            avg_standard_payload_mass = sum(standard_payload_mass_values) / len(standard_payload_mass_values)
            
            pv_payload_mass_file_name = "payload_in_kg_pv.xlsx"
            pv_payload_mass_file_path = os.path.join(item, pv_payload_mass_file_name)
            wb = load_workbook(pv_payload_mass_file_path)
            ws = wb.active
            pv_payload_mass_values = []
            for row_num in range(2, datapoint_counts + 2):
                cell_value = ws.cell(row_num, 2).value
                pv_payload_mass_values.append(cell_value)
            avg_pv_payload_mass = sum(pv_payload_mass_values) / len(pv_payload_mass_values)
            cost_of_payload_mass = round((avg_standard_payload_mass - avg_pv_payload_mass) / avg_standard_payload_mass * 100, 2)
            print(f"              The Proposed Model reduced the payload capacity (kg) by {cost_of_payload_mass}%")
            
            # Variable contains answer key of each dataset
            dataset_key = {
                "name": os.path.basename(item),
                "avg_temp": avg_temperature_value,
                "max_temp": max_temperature,
                "avg_uv": avg_uv_values,
                "max_uv": max_uv_index,
                "total_std_power_input": sum_of_standard_power_input,
                "total_std_power_consumed": sum_of_standard_power_consumptons,
                "total_pv_power_input": sum_of_pv_power_input,
                "total_pv_power_consumed": sum_of_pv_power_consumptons,
                "cost_in_power": cost_of_power_consumptions,
                "cost_in_payload_mass": cost_of_payload_mass,
                "cost_in_payload_volume": cost_of_payload_volume
            }
            
            result.append(dataset_key)
            
        # Sorting the result array...
        result = sorted(result, key=lambda item: item['name'])

        # Gnerating the answer key txt file...
        print(f"\n      Generating engr102_starship_answers_key.txt...")
        # Generating the answer folder...
        os.makedirs(os.path.join(export_directory, "Answers Key"), exist_ok=True)
        # Generating the answer key path...
        answer_key_export_path = os.path.join(export_directory, "Answers Key", "engr102_starship_answers_key.txt")
        # Begin to write the answer key txt file...
        with open(answer_key_export_path, 'w') as f: 
            f.write("ENGR 102 Starship Problem Answer Key")
            f.write("\n")
            for dataset_result in result:
                f.write(f"  {dataset_result['name']}")
                f.write("\n")
                f.write(f"\n      Avg Temperature in C: {dataset_result["avg_temp"]}°C")
                f.write(f"\n      Max Temperature in C: {dataset_result["max_temp"]}°C")
                f.write("\n")
                f.write(f"\n      Avg UV Index: {dataset_result['avg_uv']}")
                f.write(f"\n      Max UV Index: {dataset_result['max_uv']}")
                f.write("\n")
                f.write(f"\n      Sum of the Power the Standard Model Received: {dataset_result['total_std_power_input']}")
                f.write(f"\n      Sum of the Power the Standard Model consumed: {dataset_result['total_std_power_consumed']}")
                f.write("\n")
                f.write(f"\n      Sum of the Power the Proposed Model Received: {dataset_result['total_pv_power_input']}")
                f.write(f"\n      Sum of the Power the Proposed Model consumed: {dataset_result['total_pv_power_consumed']}")
                f.write("\n")
                f.write(f"\n      The Additional Power Consumed by the Proposed Model: {dataset_result['cost_in_power']} Watts")
                f.write(f"\n      The Proposed Model reduced the payload volume (m^3) by {dataset_result['cost_in_payload_volume']}%")
                f.write(f"\n      The Proposed Model reduced the payload capacity (kg) by {dataset_result['cost_in_payload_mass']}%")
                
                f.write("\n")
                f.write("\n")

def generate_random_nums(min, avg, max, counts):
    # Initialize array of nums...
    nums = []
    
    # Randomize numbers until reaching the target number of numbers...
    while len(nums) < counts:
        randomized_number = random.uniform(min, max)
        nums.append(randomized_number)
    
    # Find average of the set of randomized numbers...
    nums_average = sum(nums) / len(nums)
    # Find the difference between the target average and the randomized average...
    average_difference = round(avg - nums_average, 0) # Round to 0 deci points to add some noises to the dataset
    # Adjusted randomized numbers to reach the approximate avg...
    adjusted_nums = [num + average_difference for num in nums] 
    
    return adjusted_nums

def generate_time_stamps():
    timestamps = []
    # Set time interval where data being collected (7:30 AM - 10:00 PM)
    start = datetime.datetime.strptime('7:30:00', '%H:%M:%S')
    end = datetime.datetime.strptime('22:00:00', '%H:%M:%S')
    
    # Set delta between point of collecting...
    delta = datetime.timedelta(minutes=5)
    
    # Inititalize while loop with t...
    t = start
    while t <= end:
        timestamps.append(f"{t.strftime('%H:%M:%S')}") # Converting datetime obj to str...
        t += delta # Add 5 minutes...
    
    return timestamps
    
def generate_problem(export_directory, data):
    print("-" * 75) # seperator
    print((" " * 25) + "[1] Generating Problems" + (" " * 25))
    print("-" * 75) # seperator
    
    nums_of_datasets = len(data)
    
    for dataset in data:
        
        # The directory to the dataset that in progress of making...
        making_dataset = None
        
        while True:
            # Base on numbers of datasets in json file, random the dataset index...
            dataset_index = random.randint(1,4)
            # base on randomized dataset index, generate a folder path to that dataset...
            dataset_path = os.path.join(export_directory, f"Dataset {dataset_index}")
            
            # if the dataset directory is already exists, randomize another index and path...
            if os.path.exists(dataset_path):
                continue
            else:
                # if the dataset directory chosen is not exists...
                os.mkdir(dataset_path)
                making_dataset = dataset_path
                break
        
        print(f"\n       Generating {os.path.basename(making_dataset)}...")
        
        print(f"\n       Generated Dataset Path: {making_dataset}")
        
        # Declaring key to be ignored...
        skipped_key_value = None
        key_to_ignore = "station_power_input"
        
        # Loop through all the keys except station_power_input...
        for key in dataset:
            start_time = time.time()
            
            if key == "season":
                continue
            if key == key_to_ignore:
                skipped_key_value = key
                continue
            
            # Initializing workbook...
            wb_basename = key
            wb_path = os.path.join(making_dataset, f"{key}.xlsx")
            wb = Workbook()
            ws = wb.active
            
            # Generating (min, avg, max) of each data file...
            print(f"\n       Generating {key}.xlsx...")
            print(f"\n            (", end=" ")
            for key,value in dataset[key].items():
                print(f"{key.capitalize()}: {value}", end=" ")
            print(f")")
            
            print("\n            Generating timestamp column...")
            
            # Generating timestamp list...
            time_stamps = generate_time_stamps()
            
            # Generating timestamp column & label...
            time_stamps_column = 1
            ws.column_dimensions[get_column_letter(time_stamps_column)].width = 20
            ws.cell(1, 1, "Time (%HH:%MM:%SS)")
            
            for index, value in enumerate(time_stamps, 1):
                ws.cell(index + 1, time_stamps_column, value)
                
            print("\n            Generating values column...")
                
            # Generating randomized numbers list...
            values = generate_random_nums(dataset[wb_basename]["min"], dataset[wb_basename]["avg"], dataset[wb_basename]["max"], len(time_stamps))
            
            # Generating value column & label...
            values_column = 2
            ws.column_dimensions[get_column_letter(values_column)].width = 20
            ws.cell(1, 2, f"Values ({dataset[wb_basename]["unit"] if "unit" in dataset[wb_basename] else ""})")
            
            for index, value in enumerate(values, 1):
                ws.cell(index + 1, values_column, value)
                
            wb.save(wb_path)

            end_time = time.time()
            # set elapsed time to ms and 2 deci points...
            elapsed_time = round((end_time - start_time) * 1000, 3) 
            # print out the elapsed time...
            print(f"\n            {wb_basename}.xlsx has been generated ({elapsed_time}ms)") 
            
        # Generating station_power_input_std.xlsx & modify power_input_pv
        if skipped_key_value != None:
            start_time = time.time()
            basename = skipped_key_value
            
            # Initialize workbooks
            wb = Workbook()
            ws = wb.active
            wb_path = os.path.join(making_dataset, f"power_input_std.xlsx")
            
            # Generating (min, avg, max) of each data file...
            print(f"\n       Generating power_input_std.xlsx...")
            print(f"\n            (", end=" ")
            for key,value in dataset[basename].items():
                print(f"{key.capitalize()}: {value}", end=" ")
            print(f")")
            
            print("\n            Generating timestamp column...")
            
            # Generating timestamp list...
            time_stamps = generate_time_stamps()
            
            # Generating timestamp column & label...
            time_stamps_column = 1
            ws.cell(1, 1, "Time (%HH:%MM:%SS)")
            ws.column_dimensions[get_column_letter(time_stamps_column)].width = 20
            
            for index, value in enumerate(time_stamps, 1):
                ws.cell(index + 1, time_stamps_column, value)
                
            print("\n            Generating values column...")
                
            # Generating value column & label...
            values_column = 2
            ws.column_dimensions[get_column_letter(values_column)].width = 20
            ws.cell(1, 2, f"Values ({dataset[basename]["unit"] if "unit" in dataset[basename] else ""})")
            
            # Set values columns to 0 initially...
            current_row = 1
            while current_row <= len(time_stamps):
                ws.cell(current_row + 1, values_column, 0)
                current_row += 1
                
            # Numbers of times robo went to the charging stations...
            total_nums_of_charging_charging_intervals = random.randint(50, len(time_stamps) // 2)

            # When robot first charge of the day...
            start_interval = 0
            
            for current_nums_of_charging_intervals in range(1, total_nums_of_charging_charging_intervals + 1):
                # Randomized how long the robot will be charged...
                duration_of_intervals = random.randint(1,3)
                end_interval = start_interval + duration_of_intervals
                
                while start_interval <= end_interval:
                    # Getting value from json...
                    value = dataset[basename]["value"]
                    # Add noise
                    value = generate_random_nums(value - 0.5, value, value + 0.5, 1)[0]
                    ws.cell(start_interval + 2, values_column, value)

                    start_interval += 1
                
                start_interval = random.randint(1, len(time_stamps))
                current_nums_of_charging_intervals += 1
                
            wb.save(wb_path)
            
            end_time = time.time()
            # set elapsed time to ms and 2 deci points...
            elapsed_time = round((end_time - start_time) * 1000, 3) 
            # print out the elapsed time...
            print(f"\n            {wb_basename}.xlsx has been generated ({elapsed_time}ms)") 
        
        print("-" * 75)

def read_data_json():
    with open('data.json', 'r') as f:
        return json.load(f)["data"]

def get_folder_path(test_path="/Users/edwardl210/Documents/Coding/ENGR102_Starship_Proj/src/engr102_starship_proj"):
    print((" " * 22) + "Selecting Export Destination" + (" " * 22))
    print("-" * 75)

    # While loop until a valid path is returned...
    while True:
        if test_path:
            path = test_path
        else:
            path = input("\n       Please provide a export destination: ")

        print(f"\n       Provided Path: {path}")
        confirmation = input("\n       Is this correct: (1) Yes (2) No: ")

        if confirmation == "1":
            if os.path.isdir(path):
                print("\n       Path has been verified")
                print("")
                
                export_path = os.path.join(path, "engr102_starship_problems")
                os.makedirs(os.path.join(export_path, "Answers Key"), exist_ok=True)
                
                return export_path
            else: 
                print("       Provided path is invalid. Please try again")
                continue
        elif confirmation == '2':
            print("       Path rejected. Please try entering the correct path again.")
            continue
        else:
            print("       Invalid selection. Please enter '1' for Yes or '2' for No.")
            
def get_user_options():
    
    # While loop until users choose a valid option...
    while True:
        print("-" * 75) # Separator
        # Print the initial greeting only once
        print((" " * 10) + "Welcome to the ENGR102 Starship Problem Generator." + (" " * 10))
        print("-" * 75) # Separator
        print("\n       Options: ")
        print("         [1] Generate Problems") # Generates 4 days of data..
        print("         [2] Generate Answers Key") # Generates 4 days of data..
        print("         [3] Quit Tool\n") # Quit tools
        print("-" * 75) # Seperator
        selection = input("       Enter a menu option in the keyboard [1,2,3]: ")
        
        print("-" * 75)
        
        if selection == "1":
            return "1"
        elif selection == "2":
            return "2"
        elif selection == "3":
            return "3"
        else: 
            print("Invalid selection. Please choose one of the options from the menu")
            continue

def main():
    
    while True:
        # Get user selection...
        user_selection = get_user_options()
        if user_selection == "3":
            break
    
        # Reading data.json
        data = read_data_json()
        
        if user_selection == "1":
            # Get export destination
            export_path = get_folder_path()
            generate_problem(export_directory=export_path, data=data)
            
        if user_selection == "2":
            # Get the location of the datasets
            export_path = get_folder_path()
            generate_answer_keys(export_directory=export_path)
    
    print("-" * 75)
    for i in range(0,3):
        print(f"          Exiting program in {3-i}...") 
        time.sleep(1)
        print("-" * 75)
    exit()
    
    
    
    
    
if __name__ == "__main__":
    main()